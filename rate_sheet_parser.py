"""Parse Excel rate/price sheets into structured data for the proposal agent."""

from pathlib import Path

import openpyxl


def parse_rate_sheet(file_path: str) -> dict:
    """Parse an Excel rate or price sheet into structured data.

    Returns a dict with:
        - headers: list of column header names
        - rows: list of dicts (one per row, keyed by header)
        - raw_text: a plain-text summary suitable for LLM context
    """
    path = Path(file_path)
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    all_sheets = {}
    raw_text_parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_data = []
        headers = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            # Skip completely empty rows
            if all(cell is None for cell in row):
                continue

            if i == 0:
                headers = [str(c).strip() if c else f"Column_{j}" for j, c in enumerate(row)]
                continue

            row_dict = {}
            for j, cell in enumerate(row):
                key = headers[j] if j < len(headers) else f"Column_{j}"
                row_dict[key] = cell
            rows_data.append(row_dict)

        all_sheets[sheet_name] = {
            "headers": headers,
            "rows": rows_data,
        }

        # Build plain-text summary for LLM context
        raw_text_parts.append(f"## Sheet: {sheet_name}")
        if headers:
            raw_text_parts.append(" | ".join(headers))
            raw_text_parts.append("-" * 40)
        for row_dict in rows_data:
            vals = [str(v) if v is not None else "" for v in row_dict.values()]
            raw_text_parts.append(" | ".join(vals))

    wb.close()

    return {
        "sheets": all_sheets,
        "raw_text": "\n".join(raw_text_parts),
    }
